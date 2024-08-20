from openpyxl import Workbook, load_workbook
import random

localBombas = []

def configs():
    
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="Linhas")
        config.cell(column=1, row=2, value="Colunas")
        config.cell(column=1, row=3, value="Dificuldade")
        config.cell(column=2, row=1, value="5")
        config.cell(column=2, row=2, value="5")
        config.cell(column=2, row=3, value="1")

    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value
    dificuldade = config.cell(column=2, row=3).value

    wb.save('campo.xlsx')

    # print(str(linhas))
    # print(str(colunas))

    config = {
        'linhas'        : int(linhas),
        'colunas'       : int(colunas),
        'dificuldade'   : int(dificuldade)
    }

    return config

def CriarTabuleiro():
    config = configs()
    linhas = config['linhas']
    colunas = config['colunas']
    bombas = CalcularBombas()
    tabuleiro = []
    cont = 1

    # 0 = casa vazia | -1 = Bombas

    for i in range(0, int(linhas)):
        linha = []
        for j in range(0, int(colunas)):
            # print('{:5}'.format('[ ??? ]'), end = " ")
            if cont in bombas:
                linha.append('-1')
            else:
                linha.append('0')

            cont += 1
        # print()
        tabuleiro.append(linha)
    
    # print()
    # print(bombas)
    # print()
    # print(tabuleiro)
    # print()

    gravarTabuleiro(tabuleiro)

    # Função antiga
    # for i in range(0, int(linhas)):
    #     # for j in range(0, int(colunas)):
    #         # print('*', end = " ")
    #         # print('{:5}'.format('[ * ]'), end = " ")
    #     # print()
    #     print('[ * ] ' * int(colunas))
    #     print('------' * int(colunas))

def gravarTabuleiro(tabuleiro):
    try:
        wb = load_workbook(filename='campo.xlsx')
    except:
        wb = Workbook()

    try:
        abaJogo = wb['jogo']
    except:
        abaJogo = wb.create_sheet('jogo')

    print()

    abaJogo.delete_rows(1, abaJogo.max_row)
    abaJogo.delete_cols(1, abaJogo.max_column)

    for linha in range(1, len(tabuleiro) + 1):
        for coluna in range(1, len(tabuleiro[0]) + 1):
            abaJogo.cell(row=linha, column=coluna, value=tabuleiro[linha-1][coluna-1])
    
    print()
    wb.save('campo.xlsx')

def ConfigCampo():

    linhas = input("Digite um numero maior que 2 (linhas): \033[32m")
    print('\033[0;0m', end='')

    while int(linhas) < 3:
        linhas = input('digite um novo número que seja maior que dois (linhas): \033[32m')
        print('\033[0;0m', end='')

    colunas = input("Digite um numero maior que 2 (Colunas): \033[32m")
    print('\033[0;0m', end='')

    while int(colunas) < 3:
        colunas = input('digite um novo número que seja maior que dois (Colunas): \033[32m')
        print('\033[0;0m', end='')

    dificuldade = input("\n1 - Facil\n2 - Médio\n3 - Dificil \nEscolha uma nova dificuldade: \033[32m")
    print('\033[0;0m', end='')

    while int(dificuldade) <= 0 or int(dificuldade) >= 4:
        dificuldade = input('digite um novo número de 1 e 3: \033[32m')
        print('\033[0;0m', end='')

    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']

    except:
        wb = Workbook()
        config = wb.create_sheet('config')

    config.cell(column=1, row=1, value="Linhas")
    config.cell(column=1, row=2, value="Colunas")
    config.cell(column=1, row=3, value="dificuldade")
    config.cell(column=2, row=1, value=str(linhas))
    config.cell(column=2, row=2, value=str(colunas))
    config.cell(column=2, row=3, value=str(dificuldade))
    wb.save('campo.xlsx')

def CalcularBombas():
    '''
        Descubra o valor total de casas "casas" no tabuleiro e a partir desse total calcule a quantidade de bombas que vai ter no tabuleiro
        facil   -> 15%
        médio   -> 30%
        Dficil  -> 50%
    '''

    config = configs()
    totalCasas = config['linhas']  * config['colunas']

    if config['dificuldade'] == 1:
        porcentoBombas = 0.15

    if config['dificuldade'] == 2:
        porcentoBombas = 0.30

    if config['dificuldade'] == 3:
        porcentoBombas = 0.50

    totalBombas = int(float(totalCasas) * float(porcentoBombas))

    bombas = []

    while True:
        posicao = random.randint(1, totalCasas)

        if posicao not in bombas:
            bombas.append(posicao)

        if len(bombas) >= totalBombas:
            break

    # Função antiga
    # for i in range(0, totalBombas):
    #     numeroAleatorio = random.randint(1, totalCasas)
    #     if localBombas.count(numeroAleatorio) >= 1 :
    #         print('repetido')
    #         i = i - 1
    #     else:
    #         localBombas.append(numeroAleatorio)
    #     # print('Bomba no: ' + str(localBombas[i]))

    # localBombas.sort()
    # for i in range(0, len(localBombas)):
    #     print('Bomba no: ' + str(localBombas[i]))

    bombas.sort()
    # print(sorted(bombas))
    # print(totalBombas)
    # input()
    return bombas

def jogar():
    CriarTabuleiro()

    wb          = load_workbook(filename='campo.xlsx')
    jogo        = wb['jogo']
    maxLinha    = jogo.max_row
    maxColuna   = jogo.max_column
    gameOver    = False
    jgPossivel  = (maxLinha * maxColuna) - len(CalcularBombas())
    historico = 0

    while True:

        for linha in range(1, maxLinha+1):
            for coluna in range(1, maxColuna+1):
                casa = jogo.cell(row=linha, column=coluna).value
                if int(casa) == 0 or (int(casa) <= -1 and gameOver == False):
                    print('{:8}'.format('[ ??? ]'), end='')
                elif int(casa) == 1:
                    print('{:8}'.format('[     ]'), end='')
                elif int(casa) == -1 and gameOver == True:
                    print('{:8}'.format('[ xxx ]'), end='')
            print()
        while True:
            try:
                linhaJogada = int(input('Linha: \033[32m'))
                print('\033[0;0m', end='')

                if 1 <= linhaJogada <= maxLinha:
                    break
                else:
                    linhaJogada = int(input('Valor invalido, seleciona um novo valor: \033[32m'))
                    print('\033[0;0m', end='')


            except:
                print()
                print('\033[0;0m', end='')
                

        while True:
            try:
                colunaJogada = int(input('Coluna: \033[32m'))
                print('\033[0;0m', end='')      

                if 1 <= linhaJogada <= maxLinha:
                    break
                else:
                    colunaJogada = int(input('Valor invalido, seleciona um novo valor: \033[32m'))
                    print('\033[0;0m', end='')


            except:
                print()
                print('\033[0;0m', end='')
                

        # linhaJogada = 5
        # colunaJogada = 2

         # [ ], ⏺, 💥

        jogada = int(jogo.cell(row=linhaJogada, column=colunaJogada).value)

        if jogada == 0:
            jogo.cell(row=linhaJogada, column=colunaJogada, value=1)
            historico += 1
        elif jogada <= -1:
            # jogo.cell(row=linhaJogada, column=colunaJogada, value=-2)
            gameOver = True
        elif jogada == 1:
            print('Jogada ja efetuada, bora bill')


       
        wb.save('campo.xlsx')


        if gameOver == True:
            for linha in range(1, maxLinha+1):
                for coluna in range(1, maxColuna+1):
                    casa = jogo.cell(row=linha, column=coluna).value
                    if int(casa) == 0 or (int(casa) <= -1 and gameOver == False):
                        print('{:8}'.format('[ ??? ]'), end='')
                    elif int(casa) == 1:
                        print('{:8}'.format('[     ]'), end='')
                    elif int(casa) == -1 and gameOver == True:
                        print('{:8}'.format('[ xxx ]'), end='')
                print()
            

            print('Perdeste, vadia 🤟')
            break

        if historico == jgPossivel:
            print("Ganhaste, vadia 🤟")
            break
    op = input('Tentar de novo? (S/N): \033[32m')
    print('\033[0;0m', end='')
    if str(op).upper() == 'S':
        jogar()

    


    



CalcularBombas()
# input()
config = configs()
print('''
    _______      ____    ,---.    ,---..-------.     ,-----.            ,---.    ,---..-./`) ,---.   .--.   ____     ______         ,-----.     
   /   __  \   .'  __ `. |    \  /    |\  _(`)_ \  .'  .-,  '.          |    \  /    |\ .-.')|    \  |  | .'  __ `. |    _ `''.   .'  .-,  '.   
  | ,_/  \__) /   '  \  \|  ,  \/  ,  || (_ o._)| / ,-.|  \ _ \         |  ,  \/  ,  |/ `-' \|  ,  \ |  |/   '  \  \| _ | ) _  \ / ,-.|  \ _ \  
,-./  )       |___|  /  ||  |\_   /|  ||  (_,_) /;  \  '_ /  | :        |  |\_   /|  | `-'`"`|  |\_ \|  ||___|  /  ||( ''_'  ) |;  \  '_ /  | : 
\  '_ '`)        _.-`   ||  _( )_/ |  ||   '-.-' |  _`,/ \ _/  |        |  _( )_/ |  | .---. |  _( )_\  |   _.-`   || . (_) `. ||  _`,/ \ _/  | 
 > (_)  )  __ .'   _    || (_ o _) |  ||   |     : (  '\_/ \   ;        | (_ o _) |  | |   | | (_ o _)  |.'   _    ||(_    ._) ': (  '\_/ \   ; 
(  .  .-'_/  )|  _( )_  ||  (_,_)  |  ||   |      \ `"/  \  ) /         |  (_,_)  |  | |   | |  (_,_)\  ||  _( )_  ||  (_.\.' /  \ `"/  \  ) /  
 `-'`-'     / \ (_ o _) /|  |      |  |/   )       '. \_/``".'          |  |      |  | |   | |  |    |  |\ (_ o _) /|       .'    '. \_/``".'   
   `._____.'   '.(_,_).' '--'      '--'`---'         '-----'            '--'      '--' '---' '--'    '--' '.(_,_).' '-----'`        '-----'     
                                                                                                                                                
''')

while True:

    print('1 - Jogar (Em desenvolvimento)')
    print('2 - Configurar (Em desenvolvimento)')
    print('3 - Sair')

    seletor = input('selecione a braba: \033[32m')
    print('\033[0;0m', end='')
    

    if seletor == '1':
        # print('Jogar')
        # config = configs()
        jogar()

    elif seletor == '2':
        print('Configurar')
        ConfigCampo()

    elif seletor == '3':
        print('Saindo')
        break

    else:
        print('Opção invalida')
    
